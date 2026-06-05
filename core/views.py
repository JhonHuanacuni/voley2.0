import json
from datetime import date
from io import BytesIO
from urllib.parse import quote, urlencode

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import qrcode
except ImportError:
    qrcode = None

from .forms import PaymentForm, ShiftForm, StudentForm
from .models import Attendance, Expense, Membership, Payment, Sale, Shift, Student, active_cycle_choices, valid_cycle_id
from .receipt_pdf import fill_payment_receipt
from .attendance_matrix_export import build_attendance_matrix_workbook
from .attendance_report import (
    get_attendance_chart_stats,
    get_attendance_by_shift,
    get_monthly_enrollments,
    month_bounds,
)


def _today_iso():
    return date.today().isoformat()


def _first_of_month():
    today = date.today()
    return date(today.year, today.month, 1)


def _weekday_from_date(date_string):
    try:
        selected = date.fromisoformat(date_string)
    except (TypeError, ValueError):
        return None
    # Date.getDay equivalent: 0=domingo ... 6=sábado
    return selected.weekday() + 1 if selected.weekday() < 6 else 0


def _student_attends_on_date(student, date_string):
    return student.attends_on_weekday(_weekday_from_date(date_string))


def _shift_active_on_date(shift, date_string):
    if not shift:
        return True
    return shift.is_active_on_weekday(_weekday_from_date(date_string))


def _student_debt(student):
    paid = student.payments.aggregate(total=Sum('amount'))['total'] or 0
    if not student.monthly_fee or not student.enrollment_date:
        return {'debt': None, 'expected_total': None, 'paid_total': paid, 'months': None}
    months = (date.today().year - student.enrollment_date.year) * 12 + date.today().month - student.enrollment_date.month + 1
    expected = float(student.monthly_fee) * months
    debt = max(0, expected - float(paid))
    return {'debt': debt, 'expected_total': expected, 'paid_total': paid, 'months': months}


def get_user_role(user):
    if not user.is_authenticated:
        return None
    if user.is_superuser:
        return 'admin'
    profile = getattr(user, 'userprofile', None)
    if profile is not None:
        return profile.role
    return 'secretary'


def _ensure_admin(request):
    if get_user_role(request.user) != 'admin':
        return redirect('dashboard')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    error = ''
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        error = 'Credenciales incorrectas. Intenta de nuevo.'

    return render(request, 'core/login.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def dashboard_view(request):
    students = Student.objects.all()
    active_students = students.filter(retired=False)
    inactive_enrollment = active_students.filter(enrollment_status='inactive').count()
    attended_this_month = Attendance.objects.filter(date__gte=_first_of_month())
    payments_month = Payment.objects.filter(date__gte=_first_of_month())
    # Keep KPI aligned with memberships list (status=Deuda).
    debt_count = Membership.objects.filter(status='debt').count()
    # pagos totales del mes
    payments_month_total = payments_month.aggregate(total=Sum('amount'))['total'] or 0
    # preparar datos de los últimos 6 meses para gráfico
    from datetime import datetime
    payments_by_month_labels = []
    payments_by_month_data = []
    today = datetime.today()
    for i in range(5, -1, -1):
        m = (today.month - i - 1) % 12 + 1
        y = today.year + ((today.month - i - 1) // 12)
        label = f"{y}-{m:02d}"
        total = Payment.objects.filter(date__year=y, date__month=m).aggregate(total=Sum('amount'))['total'] or 0
        payments_by_month_labels.append(label)
        payments_by_month_data.append(float(total))

    expenses_month = Expense.objects.filter(date__gte=_first_of_month())
    expenses_month_total = expenses_month.aggregate(total=Sum('amount'))['total'] or 0
    expenses_by_month_labels = []
    expenses_by_month_data = []
    for i in range(5, -1, -1):
        m = (today.month - i - 1) % 12 + 1
        y = today.year + ((today.month - i - 1) // 12)
        label = f"{y}-{m:02d}"
        total = Expense.objects.filter(date__year=y, date__month=m).aggregate(total=Sum('amount'))['total'] or 0
        expenses_by_month_labels.append(label)
        expenses_by_month_data.append(float(total))

    today = _today_iso()
    # calcular cumpleaños hoy
    from datetime import date as _date
    today_dt = _date.today()
    birthdays_today = 0
    for s in active_students:
        if s.birth_date:
            try:
                if s.birth_date.month == today_dt.month and s.birth_date.day == today_dt.day:
                    birthdays_today += 1
            except Exception:
                pass
    user_role = get_user_role(request.user)
    is_secretary = user_role == 'secretary'
    sales_month_total = 0
    if user_role == 'admin':
        sales_month_total = Sale.objects.filter(created_at__gte=_first_of_month()).aggregate(total=Sum('price'))['total'] or 0

    att_period = request.GET.get('att_period', 'monthly')
    if att_period not in ('daily', 'weekly', 'monthly'):
        att_period = 'monthly'

    month_start_default, month_end_default = month_bounds(today_dt)
    att_start_str = request.GET.get('att_start', month_start_default.isoformat())
    att_end_str = request.GET.get('att_end', month_end_default.isoformat())
    try:
        att_start_date = date.fromisoformat(att_start_str)
    except (TypeError, ValueError):
        att_start_date = month_start_default
    try:
        att_end_date = date.fromisoformat(att_end_str)
    except (TypeError, ValueError):
        att_end_date = month_end_default
    if att_start_date > att_end_date:
        att_start_date, att_end_date = att_end_date, att_start_date

    att_shift = request.GET.get('att_shift', '')
    shift_id = int(att_shift) if str(att_shift).isdigit() else None
    attendance_charts = get_attendance_chart_stats(
        att_start_date,
        shift_id,
        today=today_dt,
    )
    monthly_enrollments = get_monthly_enrollments(att_start_date, att_end_date, shift_id)

    chart_end_date = min(att_end_date, today_dt)
    attendance_by_shift = None
    if not is_secretary:
        attendance_by_shift = get_attendance_by_shift(
            att_start_date,
            chart_end_date,
            shift_id,
        )
    if att_start_date == chart_end_date:
        att_shift_chart_period = att_start_date.strftime('%d/%m/%Y')
    else:
        att_shift_chart_period = (
            f'{att_start_date.strftime("%d/%m/%Y")} — {chart_end_date.strftime("%d/%m/%Y")}'
        )

    def _chart_stats_for_json(stats):
        return {
            **stats,
            'start_date': stats['start_date'].isoformat(),
            'end_date': stats['end_date'].isoformat(),
        }

    context = {
        'total_students': students.count(),
        'active_enrollment': active_students.filter(enrollment_status='active').count(),
        'active_count': active_students.filter(enrollment_status='active').count(),
        'inactive_enrollment': inactive_enrollment,
        'birthdays_today': birthdays_today,
        'retired_count': students.filter(retired=True).count(),
        'attendance_month_present': attended_this_month.filter(status='present').count(),
        'attendance_month_absent': attended_this_month.filter(status='absent').count(),
        'attendance_month_total': attended_this_month.count(),
        'payments_month_count': payments_month.count(),
        'payments_month_total': payments_month_total,
        'payments_total': f"{payments_month_total:.2f}",
        'payments_by_month_labels': json.dumps(payments_by_month_labels),
        'payments_by_month_data': json.dumps(payments_by_month_data),
        'expenses_month_total': expenses_month_total,
        'expenses_total': f'{expenses_month_total:.2f}',
        'sales_month_total': sales_month_total,
        'expenses_by_month_labels': json.dumps(expenses_by_month_labels),
        'expenses_by_month_data': json.dumps(expenses_by_month_data),
        'debt_count': debt_count,
        'students': students.order_by('-created_at')[:5],
        'recent_students': students.order_by('-created_at')[:5],
        'user_role': user_role,
        'is_secretary': is_secretary,
        'att_period': att_period,
        'att_start': att_start_date.isoformat(),
        'att_end': att_end_date.isoformat(),
        'att_shift': att_shift,
        'attendance_charts': attendance_charts,
        'att_chart_until_today': json.dumps(_chart_stats_for_json(attendance_charts['until_today'])),
        'monthly_enrollments': monthly_enrollments,
        'shift_choices': [(str(s.id), str(s)) for s in Shift.objects.order_by('name')],
        'att_shift_chart_period': att_shift_chart_period,
        'attendance_by_shift_chart': json.dumps(attendance_by_shift) if attendance_by_shift else 'null',
    }
    return render(request, 'core/dashboard.html', context)


def _valid_shift_filter(value):
    if not value or not str(value).isdigit():
        return ''
    if Shift.objects.filter(pk=int(value)).exists():
        return str(value)
    return ''


def _shift_choices():
    return [(str(shift.id), str(shift)) for shift in Shift.objects.order_by('name')]


STUDENT_PAGE_SIZES = (10, 20, 50)


def _student_per_page(request):
    try:
        size = int(request.GET.get('per_page', 10))
    except (TypeError, ValueError):
        return 10
    return size if size in STUDENT_PAGE_SIZES else 10


def _students_list_query_params(request, page=None, per_page=None):
    params = {}
    query = request.GET.get('q', '').strip()
    cycle_filter = valid_cycle_id(request.GET.get('cycle', ''))
    shift_filter = _valid_shift_filter(request.GET.get('shift', ''))
    size = per_page if per_page is not None else _student_per_page(request)

    if query:
        params['q'] = query
    if cycle_filter:
        params['cycle'] = cycle_filter
    if shift_filter:
        params['shift'] = shift_filter
    if size != 10:
        params['per_page'] = str(size)
    if page:
        params['page'] = str(page)
    return params


@login_required(login_url='login')
def student_list(request):
    query = request.GET.get('q', '').strip()
    cycle_filter = valid_cycle_id(request.GET.get('cycle', ''))
    shift_filter = _valid_shift_filter(request.GET.get('shift', ''))
    students = Student.objects.filter(retired=False).select_related('cycle', 'shift').prefetch_related('memberships')
    if query:
        students = students.filter(
            Q(name__icontains=query)
            | Q(contact__icontains=query)
            | Q(dni__icontains=query)
            | Q(email__icontains=query)
        )
    if cycle_filter:
        students = students.filter(cycle_id=int(cycle_filter))
    if shift_filter:
        students = students.filter(shift_id=int(shift_filter))
    students = students.order_by('name')

    per_page = _student_per_page(request)
    page_obj = Paginator(students, per_page).get_page(request.GET.get('page'))

    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('students_list')
    else:
        form = StudentForm()

    context = {
        'page_obj': page_obj,
        'per_page': per_page,
        'page_sizes': STUDENT_PAGE_SIZES,
        'students_query': urlencode(_students_list_query_params(request)),
        'query': query,
        'cycle_filter': cycle_filter,
        'cycle_choices': active_cycle_choices(),
        'shift_filter': shift_filter,
        'shift_choices': _shift_choices(),
        'form': form,
        'create_mode': True,
    }
    return render(request, 'core/students.html', context)


@login_required(login_url='login')
def student_create(request):
    form = StudentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('students_list')
    students = Student.objects.filter(retired=False).prefetch_related('memberships').order_by('name')
    per_page = _student_per_page(request)
    page_obj = Paginator(students, per_page).get_page(request.GET.get('page'))
    return render(request, 'core/students.html', {
        'page_obj': page_obj,
        'form': form,
        'create_mode': True,
        'query': '',
        'cycle_filter': '',
        'cycle_choices': active_cycle_choices(),
        'shift_filter': '',
        'shift_choices': _shift_choices(),
        'per_page': per_page,
        'page_sizes': STUDENT_PAGE_SIZES,
        'students_query': urlencode(_students_list_query_params(request)),
    })


@login_required(login_url='login')
def student_edit(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    form = StudentForm(request.POST or None, instance=student)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('students_list')
    students = Student.objects.filter(retired=False).prefetch_related('memberships').order_by('name')
    per_page = _student_per_page(request)
    page_obj = Paginator(students, per_page).get_page(request.GET.get('page'))
    return render(request, 'core/students.html', {
        'page_obj': page_obj,
        'form': form,
        'student': student,
        'edit_mode': True,
        'query': '',
        'cycle_filter': '',
        'cycle_choices': active_cycle_choices(),
        'shift_filter': '',
        'shift_choices': _shift_choices(),
        'per_page': per_page,
        'page_sizes': STUDENT_PAGE_SIZES,
        'students_query': urlencode(_students_list_query_params(request)),
    })


@login_required(login_url='login')
def student_delete(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    student.delete()
    return redirect('students_list')


@login_required(login_url='login')
def student_retire(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    if request.method == 'POST':
        student.retired = True
        student.enrollment_status = 'inactive'
        student.retired_reason = request.POST.get('reason', '')
        student.retired_at = request.POST.get('retired_at') or None
        student.save()
        return redirect('students_list')
    return render(request, 'core/student_retire.html', {'student': student, 'today': _today_iso()})


@login_required(login_url='login')
def student_reactivate(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    if request.method == 'POST':
        student.retired = False
        student.retired_reason = None
        student.retired_at = None
        student.enrollment_status = 'active'
        student.save()
    return redirect('retired_list')


@login_required(login_url='login')
def retired_list(request):
    students = Student.objects.filter(retired=True).order_by('-retired_at')
    return render(request, 'core/retired.html', {'students': students})


@login_required(login_url='login')
def shifts_list(request):
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect

    edit_shift = None
    edit_id = request.GET.get('edit')

    if request.method == 'POST':
        form = ShiftForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('shifts_list')
    elif edit_id:
        edit_shift = get_object_or_404(Shift, pk=edit_id)
        form = ShiftForm(instance=edit_shift)
    else:
        form = ShiftForm()

    shifts = Shift.objects.order_by('name')
    return render(request, 'core/shifts.html', {
        'shifts': shifts,
        'form': form,
        'create_mode': not edit_shift,
        'edit_mode': bool(edit_shift),
        'shift': edit_shift,
    })


@login_required(login_url='login')
def shift_create(request):
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect
    return redirect(f"{reverse('shifts_list')}?add=1")


@login_required(login_url='login')
def shift_edit(request, shift_id):
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect

    shift = get_object_or_404(Shift, pk=shift_id)

    if request.method == 'POST':
        form = ShiftForm(request.POST, instance=shift)
        if form.is_valid():
            form.save()
            return redirect('shifts_list')
        shifts = Shift.objects.order_by('name')
        return render(request, 'core/shifts.html', {
            'shifts': shifts,
            'form': form,
            'shift': shift,
            'edit_mode': True,
        })

    return redirect(f"{reverse('shifts_list')}?edit={shift_id}")


@login_required(login_url='login')
def shift_delete(request, shift_id):
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect

    shift = get_object_or_404(Shift, pk=shift_id)
    if request.method == 'POST':
        shift.delete()
        return redirect('shifts_list')
    return redirect('shifts_list')


@login_required(login_url='login')
def attendance_view(request):
    selected_date = request.GET.get('date', _today_iso())
    shift_filter = request.GET.get('shift', '')
    history_student = request.GET.get('history_student', '')
    history_from = request.GET.get('history_from', '')
    history_to = request.GET.get('history_to', '')

    students = Student.objects.filter(retired=False, enrollment_status='active')
    selected_shift = None
    if shift_filter:
        selected_shift = Shift.objects.filter(pk=shift_filter).first()
        students = students.filter(shift=shift_filter)
    students = students.order_by('name')
    students = [
        student for student in students
        if _student_attends_on_date(student, selected_date)
        and _shift_active_on_date(selected_shift, selected_date)
    ]

    if request.method == 'POST':
        selected_date = request.POST.get('attendance_date', _today_iso())
        shift_filter = request.POST.get('shift', shift_filter)
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status in ('present', 'absent', 'late'):
                Attendance.objects.update_or_create(
                    student=student,
                    date=selected_date,
                    defaults={'status': status},
                )
            else:
                Attendance.objects.filter(student=student, date=selected_date).delete()
        return redirect(f"{reverse('attendance')}?date={selected_date}&shift={shift_filter}")

    attendance_map = {
        att.student_id: att.status
        for att in Attendance.objects.filter(date=selected_date)
    }

    history = Attendance.objects.select_related('student').order_by('-date', 'student__name')
    if history_student:
        history = history.filter(student_id=history_student)
    if history_from:
        history = history.filter(date__gte=history_from)
    if history_to:
        history = history.filter(date__lte=history_to)
    if shift_filter:
        history = history.filter(student__shift=shift_filter)

    shift_choices = [(str(shift.id), str(shift)) for shift in Shift.objects.order_by('name')]

    return render(
        request,
        'core/attendance.html',
        {
            'students': students,
            'attendance_map': attendance_map,
            'selected_date': selected_date,
            'shift_filter': shift_filter,
            'history': history,
            'history_student': history_student,
            'history_from': history_from,
            'history_to': history_to,
            'shift_choices': shift_choices,
            'student_choices': Student.objects.filter(retired=False).order_by('name'),
        },
    )


def _parse_qr_value(qr_value):
    if not qr_value:
        return None
    qr_value = qr_value.strip()
    if qr_value.lower().startswith('student:'):
        return qr_value.split(':', 1)[1].strip()
    if qr_value.isdigit():
        return qr_value
    return None


@login_required(login_url='login')
def attendance_qr_view(request):
    status_message = ''
    status_level = 'info'

    if request.method == 'POST':
        qr_value = request.POST.get('qr_value', '').strip()
        student_id = _parse_qr_value(qr_value)
        student = Student.objects.filter(pk=student_id, retired=False, enrollment_status='active').first()
        today = date.today()

        if not student:
            status_message = 'No se encontró ninguna alumna válida para el QR escaneado.'
            status_level = 'danger'
        else:
            attendance, created = Attendance.objects.update_or_create(
                student=student,
                date=today,
                defaults={'status': 'present'},
            )
            if created:
                status_message = f'Asistencia registrada para {student.name}.'
                status_level = 'success'
            else:
                status_message = f'La asistencia de {student.name} ya estaba registrada y se actualizó a PRESENTE.'
                status_level = 'warning'

    return render(request, 'core/attendance_qr.html', {
        'status_message': status_message,
        'status_level': status_level,
    })


@login_required(login_url='login')
def student_qr_view(request, student_id):
    student = get_object_or_404(Student, pk=student_id, retired=False)
    qr_value = f'student:{student.id}'

    if qrcode is None:
        url = f'https://api.qrserver.com/v1/create-qr-code/?size=300x300&data={quote(qr_value)}'
        return HttpResponseRedirect(url)

    qr_img = qrcode.make(qr_value)
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='image/png')
    response['Content-Disposition'] = f'inline; filename="qr_student_{student.id}.png"'
    return response


@login_required(login_url='login')
def delete_attendance_record(request, attendance_id):
    attendance = get_object_or_404(Attendance, pk=attendance_id)
    if request.method == 'POST':
        attendance.delete()
    query_params = []
    for param in ['date', 'shift', 'history_student', 'history_from', 'history_to']:
        value = request.POST.get(param)
        if value:
            query_params.append(f"{param}={value}")
    redirect_url = reverse('attendance')
    if query_params:
        redirect_url += '?' + '&'.join(query_params)
    return redirect(redirect_url)


@login_required(login_url='login')
def payment_list(request):
    student_filter = request.GET.get('student', '')
    payments = Payment.objects.select_related('student').order_by('-date')
    if student_filter:
        payments = payments.filter(student_id=student_filter)
    students = Student.objects.filter(retired=False).order_by('name')

    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            redirect_url = reverse('payments')
            if student_filter:
                redirect_url += f'?student={student_filter}'
            return redirect(redirect_url)
    else:
        form = PaymentForm()

    payment_debt_line = ''
    if student_filter:
        student = Student.objects.filter(pk=student_filter).first()
        if student:
            paid_total = Payment.objects.filter(student=student).aggregate(total=Sum('amount'))['total'] or 0
            debt_info = _student_debt(student)
            if debt_info['debt'] is not None:
                payment_debt_line = f"Total abonado: S/ {paid_total:.2f}. Deuda estimada: S/ {debt_info['debt']:.2f}."
            else:
                payment_debt_line = f"Total abonado: S/ {paid_total:.2f}. Configura cuota e inscripción para estimar deuda."

    return render(
        request,
        'core/payments.html',
        {
            'payments': payments,
            'students': students,
            'selected_student': student_filter,
            'form': form,
            'create_mode': True,
            'payment_debt_line': payment_debt_line,
        },
    )


@login_required(login_url='login')
def payment_create(request):
    form = PaymentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('payments')
    students = Student.objects.filter(retired=False).order_by('name')
    return render(
        request,
        'core/payments.html',
        {'payments': Payment.objects.select_related('student').order_by('-date'), 'students': students, 'form': form, 'create_mode': True},
    )


@login_required(login_url='login')
def payment_edit(request, payment_id):
    payment = get_object_or_404(Payment, pk=payment_id)
    form = PaymentForm(request.POST or None, instance=payment)
    student_filter = request.GET.get('student', '') or str(payment.student_id)
    if request.method == 'POST' and form.is_valid():
        form.save()
        redirect_url = reverse('payments')
        if student_filter:
            redirect_url += f'?student={student_filter}'
        return redirect(redirect_url)
    students = Student.objects.filter(retired=False).order_by('name')
    payments = Payment.objects.select_related('student').order_by('-date')
    if student_filter:
        payments = payments.filter(student_id=student_filter)
    paid_total = Payment.objects.filter(student=payment.student).aggregate(total=Sum('amount'))['total'] or 0
    debt_info = _student_debt(payment.student)
    if debt_info['debt'] is not None:
        payment_debt_line = f"Total abonado: S/ {paid_total:.2f}. Deuda estimada: S/ {debt_info['debt']:.2f}."
    else:
        payment_debt_line = f"Total abonado: S/ {paid_total:.2f}. Configura cuota e inscripción para estimar deuda."
    return render(
        request,
        'core/payments.html',
        {
            'payments': payments,
            'students': students,
            'form': form,
            'payment': payment,
            'edit_mode': True,
            'selected_student': student_filter,
            'payment_debt_line': payment_debt_line,
        },
    )


@login_required(login_url='login')
def payment_delete(request, payment_id):
    payment = get_object_or_404(Payment, pk=payment_id)
    payment.delete()
    return redirect('payments')


@login_required(login_url='login')
def payment_edit_redirect(request, payment_id):
    return redirect('membership_payment_edit', payment_id=payment_id)


@login_required(login_url='login')
def payment_delete_redirect(request, payment_id):
    payment = get_object_or_404(Payment, pk=payment_id)
    if get_user_role(request.user) == 'secretary':
        return redirect('membership_payments_list')
    membership_id = payment.membership_id
    payment.delete()
    if membership_id:
        from .models import Membership
        try:
            Membership.objects.get(pk=membership_id).recalculate_status()
        except Membership.DoesNotExist:
            pass
    if membership_id:
        return redirect('membership_payments', membership_id=membership_id)
    return redirect('membership_payments_list')


@login_required(login_url='login')
def payment_receipt(request, payment_id):
    payment = get_object_or_404(Payment, pk=payment_id)
    student = payment.student

    months_es = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
    }
    month_name = months_es.get(payment.date.month, 'Desconocido')

    buffer = fill_payment_receipt(payment, student, month_name)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="recibo_{payment.id}.pdf"'
    return response


@login_required(login_url='login')
def report_view(request):
    selected_student = request.GET.get('student', '')
    selected_shift = request.GET.get('shift', '')
    today = date.today()
    try:
        report_year = int(request.GET.get('year', today.year))
    except (TypeError, ValueError):
        report_year = today.year
    try:
        report_month = int(request.GET.get('month', today.month))
    except (TypeError, ValueError):
        report_month = today.month
    if report_month < 1 or report_month > 12:
        report_month = today.month

    students = Student.objects.order_by('name')
    shift_choices = [(str(shift.id), str(shift)) for shift in Shift.objects.order_by('name')]
    return render(request, 'core/reports.html', {
        'students': students,
        'shift_choices': shift_choices,
        'selected_student': selected_student,
        'selected_shift': selected_shift,
        'report_year': report_year,
        'report_month': report_month,
    })


def _write_workbook(rows, headers, money_cols=()):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col_index in money_cols:
        for row in ws.iter_rows(min_row=2, min_col=col_index + 1, max_col=col_index + 1):
            for cell in row:
                cell.number_format = '#,##0.00'
    return wb


def _write_professional_workbook(rows, headers, title, subtitle, money_cols=()):
    """Create a professional-looking Excel workbook with title, styling, and footer."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'
    
    # Set column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        ws.column_dimensions[col_letter].width = 18
    
    # Row 1: Title/Logo
    ws.merge_cells('A1:' + chr(64 + len(headers)) + '1')
    title_cell = ws['A1']
    title_cell.value = '🏐 VITA VOLEY ACADEMIA'
    title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Row 2: Subtitle
    ws.merge_cells('A2:' + chr(64 + len(headers)) + '2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = subtitle
    subtitle_cell.font = Font(name='Calibri', size=12, italic=True, color='404040')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    
    # Row 3: Empty spacer
    ws.row_dimensions[3].height = 8
    
    # Row 4: Column Headers with styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 25
    
    # Data rows
    data_font = Font(name='Calibri', size=10)
    
    for row_idx, row_data in enumerate(rows, 5):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Format money columns
            if (col_num - 1) in money_cols:
                cell.number_format = 'S/ #,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Alternate row colors for better readability
        if row_idx % 2 == 0:
            light_fill = PatternFill(start_color='E7F0F7', end_color='E7F0F7', fill_type='solid')
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_num).fill = light_fill
    
    # Footer section
    footer_row = len(rows) + 6
    ws.merge_cells(f'A{footer_row}:' + chr(64 + len(headers)) + str(footer_row))
    footer_cell = ws[f'A{footer_row}']
    footer_cell.value = 'Reporte generado por el Sistema de Gestión VITA VOLEY'
    footer_cell.font = Font(name='Calibri', size=9, italic=True, color='808080')
    footer_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    return wb


@login_required(login_url='login')
def export_students_xlsx(request):
    student_id = request.GET.get('student')
    students = Student.objects.order_by('name')
    if student_id:
        students = students.filter(pk=student_id)
    rows = []
    for student in students:
        rows.append([
            student.name,
            student.age or '',
            student.dni or '',
            student.email or '',
            student.contact or '',
            student.get_shift_display(),
            student.get_enrollment_status_display(),
            student.monthly_fee or '',
            student.enrollment_date or '',
            student.membership_start or '',
            student.membership_end or '',
            student.retired_at or '',
            student.attendance_days_display,
        ])
    wb = _write_professional_workbook(
        rows,
        [
            'Nombre',
            'Edad',
            'DNI',
            'Email',
            'Contacto',
            'Turno',
            'Matrícula',
            'Cuota mensual',
            'Inscripción',
            'Inicio membresía',
            'Fin membresía',
            'Retirada',
            'Días de asistencia',
        ],
        title='Listado de Alumnas',
        subtitle='Registro de Alumnas y Detalles de Matrícula',
        money_cols=(7,),
    )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vita_voley_alumnas.xlsx'
    wb.save(response)
    return response


@login_required(login_url='login')
def export_attendance_xlsx(request):
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
    except (TypeError, ValueError):
        year = today.year
    try:
        month = int(request.GET.get('month', today.month))
    except (TypeError, ValueError):
        month = today.month
    if month < 1 or month > 12:
        month = today.month

    student_id = request.GET.get('student') or None
    shift_id = request.GET.get('shift') or None
    include_membership_payment = get_user_role(request.user) == 'admin'

    wb = build_attendance_matrix_workbook(
        year,
        month,
        student_id,
        shift_id,
        include_membership_payment=include_membership_payment,
    )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename=vita_voley_asistencia_{year}_{month:02d}.xlsx'
    )
    wb.save(response)
    return response


def _parse_dashboard_date_range(request):
    today_dt = date.today()
    month_start_default, month_end_default = month_bounds(today_dt)
    try:
        att_start_date = date.fromisoformat(request.GET.get('att_start', month_start_default.isoformat()))
    except (TypeError, ValueError):
        att_start_date = month_start_default
    try:
        att_end_date = date.fromisoformat(request.GET.get('att_end', month_end_default.isoformat()))
    except (TypeError, ValueError):
        att_end_date = month_end_default
    if att_start_date > att_end_date:
        att_start_date, att_end_date = att_end_date, att_start_date
    return att_start_date, att_end_date


@login_required(login_url='login')
def export_monthly_enrollments_xlsx(request):
    att_start_date, att_end_date = _parse_dashboard_date_range(request)
    shift_id = request.GET.get('att_shift') or None

    enrollment_data = get_monthly_enrollments(att_start_date, att_end_date, shift_id)

    shift_label = ''
    if shift_id:
        shift = Shift.objects.filter(pk=shift_id).first()
        if shift:
            shift_label = f' | Turno: {shift.name}'

    subtitle = (
        f'Matriculados del periodo: {att_start_date.strftime("%d/%m/%Y")} — '
        f'{att_end_date.strftime("%d/%m/%Y")} | '
        f'Total: {enrollment_data["total"]} '
        f'({enrollment_data["active_count"]} activas){shift_label}'
    )

    rows = []
    for index, student in enumerate(enrollment_data['students'], start=1):
        rows.append([
            index,
            student.name,
            student.dni or '',
            student.email or '',
            student.contact or '',
            student.guardian or '',
            student.get_shift_display(),
            student.attendance_days_display,
            student.enrollment_date.strftime('%d/%m/%Y') if student.enrollment_date else '',
            student.membership_start.strftime('%d/%m/%Y') if student.membership_start else '',
            student.membership_end.strftime('%d/%m/%Y') if student.membership_end else '',
            float(student.monthly_fee) if student.monthly_fee else '',
            student.get_enrollment_status_display(),
        ])

    wb = _write_professional_workbook(
        rows,
        [
            'N°',
            'Nombre',
            'DNI',
            'Email',
            'Teléfono',
            'Apoderado',
            'Turno',
            'Horario',
            'Fecha inscripción',
            'Inicio membresía',
            'Fin membresía',
            'Cuota mensual',
            'Estado matrícula',
        ],
        title='Matriculados del periodo',
        subtitle=subtitle,
        money_cols=(11,),
    )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename=vita_voley_matriculados_{att_start_date:%Y%m%d}_{att_end_date:%Y%m%d}.xlsx'
    )
    wb.save(response)
    return response


@login_required(login_url='login')
def export_payments_xlsx(request):
    if get_user_role(request.user) == 'secretary':
        return redirect('reports')
    student_id = request.GET.get('student')
    payments = Payment.objects.select_related('membership', 'student').order_by('-date')
    if student_id:
        payments = payments.filter(student_id=student_id)
    rows = []
    for payment in payments:
        period = ''
        if payment.membership:
            period = f'{payment.membership.start_date} — {payment.membership.end_date}'
        rows.append([
            payment.student.name,
            period,
            payment.date,
            float(payment.amount),
            payment.get_method_display(),
            payment.membership.get_status_display() if payment.membership else '',
        ])
    wb = _write_professional_workbook(
        rows,
        ['Alumna', 'Periodo membresía', 'Fecha pago', 'Monto', 'Método', 'Estado membresía'],
        title='Historial de Pagos',
        subtitle='Historial de Transacciones y Pagos',
        money_cols=(3,),
    )
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vita_voley_pagos.xlsx'
    wb.save(response)
    return response
