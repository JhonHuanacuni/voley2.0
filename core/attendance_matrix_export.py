"""Exportación Excel: matriz mensual de asistencia (formato planilla escolar)."""

import calendar
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.db.models import Prefetch

from .models import Attendance, Membership, Student

WEEKDAY_ABBR_ES = ['lun', 'mar', 'mié', 'jue', 'vie', 'sáb', 'dom']

STATUS_CODE = {
    'present': 'A',
    'late': 'T',
    'absent': 'F',
}

HEADER_FILL = PatternFill(start_color='4A86E8', end_color='4A86E8', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
SUNDAY_HEADER_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
SUNDAY_HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='000000')
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_ABSENT = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
FONT_ABSENT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
FILL_MEMBERSHIP_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_MEMBERSHIP_DEBT = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
MEMBERSHIP_COL_WIDTH = 26
DATA_FONT = Font(name='Calibri', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)


def _weekday_from_date(target_date):
    # 0=domingo ... 6=sábado (mismo criterio que views.py)
    return target_date.weekday() + 1 if target_date.weekday() < 6 else 0


def _student_attends_on_date(student, target_date):
    weekday = _weekday_from_date(target_date)
    days = student.attendance_days or []
    if not days or len(days) == 7:
        return True
    return weekday in days


def _shift_active_on_date(shift, target_date):
    if not shift:
        return True
    return shift.is_active_on_weekday(_weekday_from_date(target_date))


def _enrollment_label(student):
    if student.retired:
        return 'RETIRADA'
    if student.enrollment_status == 'active':
        return 'ACTIVO'
    return 'INACTIVO'


def _day_header(target_date):
    abbr = WEEKDAY_ABBR_ES[target_date.weekday()]
    return f'{abbr} {target_date.day:02d}'


def _membership_for_month(student, month_start, month_end):
    for membership in student.memberships.all():
        if membership.start_date <= month_end and membership.end_date >= month_start:
            return membership
    memberships = student.memberships.all()
    return memberships[0] if memberships else None


def _membership_payment_cell(student, month_start, month_end):
    membership = _membership_for_month(student, month_start, month_end)
    if not membership:
        return 'Sin membresía', None

    due = float(membership.amount_due)
    if due <= 0:
        return 'No se colocó monto de membresía', None

    expiry_str = membership.end_date.strftime('%d/%m/%Y')
    header = f'TOTAL: S/ {due:.2f} - {expiry_str}'

    all_payments = sorted(membership.payments.all(), key=lambda payment: payment.date)
    payment_lines = [
        f'S/ {float(payment.amount):.2f} - {payment.date.strftime("%d/%m/%Y")}'
        for payment in all_payments
    ]

    paid_total = sum(float(payment.amount) for payment in all_payments)
    balance = max(due - paid_total, 0)
    is_complete = balance <= 0

    lines = [header, *payment_lines]
    if not is_complete:
        lines.append(f'Pendiente: S/ {balance:.2f}')

    return '\n'.join(lines), is_complete


def build_attendance_matrix_workbook(
    year,
    month,
    student_id=None,
    shift_id=None,
    include_membership_payment=False,
):
    _, days_in_month = calendar.monthrange(year, month)
    month_start = date(year, month, 1)
    month_end = date(year, month, days_in_month)

    membership_prefetch = Prefetch(
        'memberships',
        queryset=Membership.objects.order_by('-end_date', '-created_at').prefetch_related('payments'),
    )
    students = (
        Student.objects.filter(retired=False)
        .select_related('shift')
        .prefetch_related(membership_prefetch)
        .order_by('name')
    )
    if student_id:
        students = students.filter(pk=student_id)
    if shift_id:
        students = students.filter(shift_id=shift_id)
    students = list(students)

    attendance_map = {}
    attendance_qs = Attendance.objects.filter(
        date__gte=month_start,
        date__lte=month_end,
        student_id__in=[student.pk for student in students],
    )
    for record in attendance_qs:
        attendance_map[(record.student_id, record.date)] = record.status

    fixed_headers = [
        'N°',
        'NOMBRES Y APELLIDOS',
        'APODERADO',
        'TURNO',
        'HORARIO',
        'ESTADO',
    ]
    if include_membership_payment:
        fixed_headers.append('PAGO\nMEMBRESÍA')
    day_headers = [_day_header(date(year, month, day)) for day in range(1, days_in_month + 1)]
    tail_headers = ['', 'TOTAL\nASIST', 'TOTAL\nTARD', 'TOTAL\nFALTAS']
    headers = fixed_headers + day_headers + tail_headers

    wb = Workbook()
    ws = wb.active
    ws.title = 'Asistencia'

    first_day_col = len(fixed_headers) + 1
    last_day_col = first_day_col + days_in_month - 1
    membership_col = len(fixed_headers) if include_membership_payment else None
    total_asist_col = last_day_col + 2
    total_tard_col = last_day_col + 3
    total_faltas_col = last_day_col + 4

    sunday_cols = {
        first_day_col + day - 1
        for day in range(1, days_in_month + 1)
        if date(year, month, day).weekday() == 6
    }

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        is_sunday_col = col_idx in sunday_cols
        cell.font = SUNDAY_HEADER_FONT if is_sunday_col else HEADER_FONT
        cell.fill = SUNDAY_HEADER_FILL if is_sunday_col else HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if first_day_col <= col_idx <= last_day_col:
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                text_rotation=45,
                wrap_text=True,
            )

    ws.row_dimensions[1].height = 52

    for row_idx, student in enumerate(students, start=2):
        shift = student.shift
        row_values = [
            row_idx - 1,
            student.name.upper(),
            (student.guardian or '').upper(),
            (shift.name if shift else ''),
            (shift.schedule if shift else 'VOLEY VITA'),
            _enrollment_label(student),
        ]
        membership_payment_state = None
        if include_membership_payment:
            payment_text, membership_payment_state = _membership_payment_cell(
                student,
                month_start,
                month_end,
            )
            row_values.append(payment_text)

        total_asist = 0
        total_tard = 0
        total_faltas = 0

        for day in range(1, days_in_month + 1):
            current_date = date(year, month, day)
            code = ''

            # Domingos siempre en blanco (como la plantilla de referencia).
            if current_date.weekday() != 6:
                status = attendance_map.get((student.pk, current_date))
                if status:
                    # Si hay registro guardado (presente, ausente o tarde), mostrarlo siempre.
                    code = STATUS_CODE.get(status, '')
                elif (
                    _student_attends_on_date(student, current_date)
                    and _shift_active_on_date(shift, current_date)
                ):
                    code = ''

                if code == 'A':
                    total_asist += 1
                elif code == 'T':
                    total_tard += 1
                elif code == 'F':
                    total_faltas += 1

            row_values.append(code)

        row_values.extend(['', total_asist, total_tard, total_faltas])

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif first_day_col <= col_idx <= last_day_col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx in sunday_cols:
                    cell.fill = FILL_WHITE
                elif value == 'F':
                    cell.fill = FILL_ABSENT
                    cell.font = FONT_ABSENT
            elif membership_col and col_idx == membership_col:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                if membership_payment_state is True:
                    cell.fill = FILL_MEMBERSHIP_OK
                elif membership_payment_state is False:
                    cell.fill = FILL_MEMBERSHIP_DEBT
            elif col_idx in (total_asist_col, total_tard_col, total_faltas_col):
                cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 10
    if include_membership_payment and membership_col:
        ws.column_dimensions[get_column_letter(membership_col)].width = MEMBERSHIP_COL_WIDTH
    for col_idx in range(first_day_col, last_day_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4.5
    ws.column_dimensions[get_column_letter(last_day_col + 1)].width = 2
    ws.column_dimensions[get_column_letter(total_asist_col)].width = 10
    ws.column_dimensions[get_column_letter(total_tard_col)].width = 10
    ws.column_dimensions[get_column_letter(total_faltas_col)].width = 11

    ws.freeze_panes = ws.cell(row=2, column=first_day_col)

    return wb
