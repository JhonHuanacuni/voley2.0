"""Exportación Excel: reporte de ingresos, egresos y utilidad."""

import calendar
from datetime import date, timedelta

from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Expense, Payment, Sale
from .time_utils import local_datetime_range

HEADER_FILL = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='D9D2E9', end_color='D9D2E9', fill_type='solid')
SUMMARY_FILL = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
TITLE_FONT = Font(name='Calibri', size=14, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)

SALE_HEADERS = [
    'Fecha',
    'Concepto',
    'Proveedor',
    'Monto (S/.)',
    'Medio de Pago',
    'Observaciones',
]

SUMMARY_HEADERS = ['MES ANTERIOR', 'INGRESOS', 'EGRESOS', 'UTILIDAD']

MONTHS_ES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}


def _month_bounds(year, month):
    _, last_day = calendar.monthrange(year, month)
    return date(year, month, 1), date(year, month, last_day)


def _sum_payments(start_date, end_date):
    return float(
        Payment.objects.filter(date__gte=start_date, date__lte=end_date)
        .aggregate(total=Sum('amount'))['total'] or 0
    )


def _sum_sales(start_date, end_date):
    range_start, range_end = local_datetime_range(start_date, end_date)
    return float(
        Sale.objects.filter(created_at__gte=range_start, created_at__lte=range_end)
        .aggregate(total=Sum('price'))['total'] or 0
    )


def _sum_expenses(start_date, end_date):
    return float(
        Expense.objects.filter(date__gte=start_date, date__lte=end_date)
        .aggregate(total=Sum('amount'))['total'] or 0
    )


def compute_financial_summary(year, month):
    month_start, month_end = _month_bounds(year, month)
    day_before = month_start - timedelta(days=1)

    income_before = _sum_payments(date(2000, 1, 1), day_before) + _sum_sales(date(2000, 1, 1), day_before)
    expenses_before = _sum_expenses(date(2000, 1, 1), day_before)
    mes_anterior = income_before - expenses_before

    payments_total = _sum_payments(month_start, month_end)
    sales_total = _sum_sales(month_start, month_end)
    ingresos = payments_total + sales_total
    egresos = _sum_expenses(month_start, month_end)
    utilidad = mes_anterior + ingresos - egresos

    expenses = (
        Expense.objects.filter(date__gte=month_start, date__lte=month_end)
        .order_by('date', 'concept')
    )

    return {
        'year': year,
        'month': month,
        'month_label': MONTHS_ES.get(month, str(month)),
        'month_start': month_start,
        'month_end': month_end,
        'expenses': list(expenses),
        'sales_total': sales_total,
        'payments_total': payments_total,
        'ingresos': ingresos,
        'egresos': egresos,
        'mes_anterior': mes_anterior,
        'utilidad': utilidad,
    }


def _style_cell(cell, *, fill=None, bold=False, align='left', number_format=None):
    cell.font = Font(name='Calibri', size=10, bold=bold)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format


def build_financial_report_workbook(year, month):
    data = compute_financial_summary(year, month)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte financiero'

    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = (
        f'VITA VOLEY — Reporte de ingresos y egresos — '
        f'{data["month_label"]} {year}'
    )
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:F2')
    ws['A2'] = (
        f'Periodo: {data["month_start"].strftime("%d/%m/%Y")} — '
        f'{data["month_end"].strftime("%d/%m/%Y")}'
    )
    ws['A2'].alignment = Alignment(horizontal='center')

    header_row = 4
    for col_idx, header in enumerate(SALE_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        _style_cell(cell, fill=HEADER_FILL, bold=True, align='center')

    row_idx = header_row + 1
    for expense in data['expenses']:
        row_values = [
            expense.date.strftime('%d/%m/%Y'),
            expense.concept,
            expense.provider or '',
            float(expense.amount),
            expense.get_payment_method_display(),
            expense.observations or '',
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            number_format = '#,##0.00' if col_idx == 4 else None
            _style_cell(cell, number_format=number_format)
        row_idx += 1

    if not data['expenses']:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        cell = ws.cell(row=row_idx, column=1, value='Sin egresos en este periodo')
        _style_cell(cell, align='center')
        row_idx += 1

    for col_idx in range(1, 4):
        cell = ws.cell(row=row_idx, column=col_idx, value='TOTAL' if col_idx == 1 else '')
        _style_cell(cell, fill=TOTAL_FILL, bold=True, align='right' if col_idx == 1 else 'center')
    total_cell = ws.cell(row=row_idx, column=4, value=data['egresos'])
    _style_cell(total_cell, fill=TOTAL_FILL, bold=True, number_format='#,##0.00')
    for col_idx in range(5, 7):
        _style_cell(ws.cell(row=row_idx, column=col_idx, value=''), fill=TOTAL_FILL)

    summary_header_row = row_idx + 2
    for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=summary_header_row, column=col_idx, value=header)
        _style_cell(cell, fill=SUMMARY_FILL, bold=True, align='center')

    summary_values = [
        data['mes_anterior'],
        data['ingresos'],
        data['egresos'],
        data['utilidad'],
    ]
    summary_row = summary_header_row + 1
    for col_idx, value in enumerate(summary_values, start=1):
        cell = ws.cell(row=summary_row, column=col_idx, value=value)
        _style_cell(cell, fill=SUMMARY_FILL, bold=True, align='center', number_format='#,##0.00')

    notes_row = summary_row + 2
    ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=6)
    ws.cell(
        row=notes_row,
        column=1,
        value=(
            f'Ingresos del mes: pagos de membresía S/ {data["payments_total"]:.2f} + '
            f'ventas S/ {data["sales_total"]:.2f}. '
            f'Utilidad = Mes anterior + Ingresos - Egresos.'
        ),
    )
    ws[f'A{notes_row}'].font = Font(name='Calibri', size=9, italic=True)
    ws[f'A{notes_row}'].alignment = Alignment(wrap_text=True)

    widths = [12, 28, 22, 14, 16, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    return wb, data
